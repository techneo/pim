from django.http import HttpResponse

def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"Name"),
        smart_str(u"Email"),
        smart_str(u"Phone Number"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.name),
            smart_str(obj.email),
            smart_str(obj.phone),
        ])
    return response
export_csv.short_description = u"Export CSV"

def export_txt(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename=mymodel.txt'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str(u"Name"),
        smart_str(u"Email"),
        smart_str(u"Phone Number"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.name),
            smart_str(obj.email),
            smart_str(obj.phone),
        ])
    return response
export_txt.short_description = u"Export TEXT"

#pip install xlwt
def export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("MyModel")
    
    row_num = 0
    
    columns = [
        (u"Name", 2000),
        (u"Email", 6000),
        (u"Phone", 8000),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in xrange(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    
    for obj in queryset:
        row_num += 1
        row = [
            obj.name,
            obj.email,
            obj.phone,
        ]
        for col_num in xrange(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
            
    wb.save(response)
    return response
    
export_xls.short_description = u"Export XLS"

#pip install openpyxl
def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        (u"Name", 15),
        (u"Email", 70),
        (u"Phone", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.name,
            obj.email,
            obj.phone,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"